# ==============================================================================
# VIBRIS INDUSTRIAL HMI - 1600+ LINES ENTERPRISE ENGINE (UNIFIED DUAL-MODE HMI)
# Direktif: Peer-to-Peer, Brutal Technical Honesty, Zero Sycophancy.
# ==============================================================================

import sys
import os
import csv
import json
import time
from datetime import datetime
from collections import deque

try:
    import serial
    import serial.tools.list_ports
    import threading
except ImportError:
    serial = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QHBoxLayout, QVBoxLayout,
    QGridLayout, QStackedWidget, QFrame, QListWidget, QComboBox, QMessageBox,
    QSlider, QSizePolicy, QProgressBar, QLineEdit, QTableWidget, QTableWidgetItem
)
from PyQt5.QtCore import QTimer, Qt, QPoint, QEvent
from PyQt5.QtGui import QFont, QPainter, QLinearGradient, QColor, QPolygon, QPen, QBrush
import pyqtgraph as pg

# ===================== KONFIGURASI OPERASIONAL =====================
SERIAL_PORT = '/dev/ttyACM0'
BAUD_RATE = 115200
LOG_DIR = "logs"
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

ESP32_USB_HINTS = ["CH340", "CH343", "CP210", "USB-SERIAL", "USB SERIAL", "FTDI", "SILICON LABS", "COM3", "ACM0"]

# ===================== PALET WARNA INDUSTRI VIBRIS =====================
COL_BG_MAIN = "#181b20"
COL_PANEL_DARK = "#101216"
COL_ACCENT = "#38bdf8"
COL_ACCENT_DIM = "#1e2733"
COL_TEXT_LIGHT = "#e8eaed"
COL_TEXT_DIM = "#7c8592"
COL_OK = "#34d399"
COL_WARN = "#fbbf24"
COL_BAD = "#f87171"
COL_IDLE = "#94a3b8"
COL_HEADER_BG = "#1c2027"

D2_THRESHOLD_WASPADA = 9.49
D2_THRESHOLD_BAHAYA = 13.28

STATUS_SEVERITY = {"diam": -1, "normal": 0, "waspada": 1, "bahaya": 2}

# ===================== KOMPONEN KUSTOM GAUGE & GRAFIK =====================
class _GaugeBar(QWidget):
    def __init__(self, min_val, max_val):
        super().__init__()
        self.min_val = min_val
        self.max_val = max_val
        self.val = min_val

    def set_value(self, val):
        self.val = max(self.min_val, min(self.max_val, val))
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        w = self.width()
        h = self.height() - 4

        grad = QLinearGradient(0, 0, w, 0)
        grad.setColorAt(0.0, QColor(COL_OK))
        grad.setColorAt(0.5, QColor(COL_WARN))
        grad.setColorAt(1.0, QColor(COL_BAD))

        painter.setBrush(QBrush(grad))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(0, 0, w, h, 2, 2)

        ratio = (self.val - self.min_val) / (self.max_val - self.min_val) if self.max_val > self.min_val else 0
        px = int(ratio * w)

        poly = QPolygon([
            QPoint(px, h),
            QPoint(max(0, px - 4), h + 4),
            QPoint(min(w, px + 4), h + 4)
        ])
        painter.setBrush(QBrush(QColor("#ffffff")))
        painter.setPen(QPen(QColor("#ffffff"), 1))
        painter.drawPolygon(poly)

class GradientGauge(QWidget):
    def __init__(self, title, min_val, max_val, t_warn, t_danger, unit=""):
        super().__init__()
        self.min_val = min_val
        self.max_val = max_val
        self.t_warn = t_warn
        self.t_danger = t_danger
        self.unit = unit
        self.current_val = min_val

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(1)

        self.lbl_title = QLabel(title)
        self.lbl_title.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        layout.addWidget(self.lbl_title)

        self.bar = _GaugeBar(min_val, max_val)
        self.bar.setFixedHeight(16)
        layout.addWidget(self.bar)

        self.lbl_status = QLabel("Normal")
        self.lbl_status.setStyleSheet(f"font-size: 8px; color: {COL_TEXT_DIM};")
        self.lbl_status.setWordWrap(True)
        layout.addWidget(self.lbl_status)

    def set_value(self, val, status_key="normal"):
        if val is None: return
        self.current_val = val
        self.bar.set_value(val)

        if status_key == "diam":
            status, desc = "Diam / Off", "Mati"
            col = COL_IDLE
        elif val >= self.t_danger:
            status, desc = "Extreme", "Bahaya"
            col = COL_BAD
        elif val >= self.t_warn:
            status, desc = "Moderate", "Waspada"
            col = COL_WARN
        else:
            status, desc = "Normal", "Aman"
            col = COL_OK

        self.lbl_status.setText(f"<span style='color:{col}; font-weight:bold;'>{status}</span>")

# ===================== KELAS UTAMA DASHBOARD =====================
class Dashboard(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("HMI | VIBRIS PIMNAS - 1600+ ENTERPRISE ENGINE (UNIFIED DUAL-MODE)")
        self.setWindowFlags(
            Qt.Window | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint
        )
        self.setFixedSize(480, 320)
        self.setStyleSheet(f"background-color: {COL_BG_MAIN}; color: {COL_TEXT_LIGHT}; font-family: Arial;")

        screen = QApplication.primaryScreen()
        if screen is not None:
            screen_geo = screen.geometry()
            self.move(screen_geo.x(), screen_geo.y())

        self.data_len = 50
        self.time_buffer = deque(maxlen=self.data_len)
        self.v_buffer = deque(maxlen=self.data_len)
        self.vx_buffer = deque(maxlen=self.data_len)
        self.vy_buffer = deque(maxlen=self.data_len)
        self.vz_buffer = deque(maxlen=self.data_len)
        self.a_buffer = deque(maxlen=self.data_len)
        self.temp_buffer = deque(maxlen=self.data_len)
        self.rpm_buffer = deque(maxlen=self.data_len)
        self.d2_buffer = deque(maxlen=self.data_len)
        
        self.history_window_size = 100
        self.hist_v_buf = deque(maxlen=self.history_window_size)
        self.hist_a_buf = deque(maxlen=self.history_window_size)
        self.hist_temp_buf = deque(maxlen=self.history_window_size)
        self.hist_vx_buf = deque(maxlen=self.history_window_size)
        self.hist_vy_buf = deque(maxlen=self.history_window_size)
        self.hist_vz_buf = deque(maxlen=self.history_window_size)
        self.hist_rpm_buf = deque(maxlen=self.history_window_size)
        self.hist_d2_buf = deque(maxlen=self.history_window_size)

        self.tick = 0
        self.last_processed_tick = 0
        self.last_raw_line = ""
        self.last_packet_time = time.time()
        self.packet_loss_flag = False
        self.last_auto_snapshot_status = "normal"

        self.fft_hz_buffer = []
        self.fft_mag_buffer = []

        self.current_health_score = 0.0
        self.current_trend = "Menunggu"
        self.current_servis = "Menunggu"
        self.current_ml_label = "N/A"
        self.current_ml_conf = 0.0
        self.current_diag_label = "N/A"
        self.current_diag_conf = 0.0
        self.current_kurtosis = 0.0
        self.current_diagnosis_flags = "N/A"

        self.logdet_times = []
        self.logdet_v_vals = []
        self.logdet_a_vals = []
        self.logdet_t_vals = []
        self.logdet_play_index = 0
        self.logdet_play_speed = 1.0
        self.logdet_timer = QTimer(self)
        self.logdet_timer.timeout.connect(self._logdet_step)

        self.current_v = None
        self.current_vx = None
        self.current_vy = None
        self.current_vz = None
        self.current_a = None
        self.current_temp = None
        self.current_rpm = None
        self.current_d2 = None
        self.current_status_device = ""

        self.session_sample_count = 0
        self.session_rpm_sum = 0.0
        self.session_d2_max = 0.0
        self.session_worst_status = "Normal"
        self.session_waspada_count = 0
        self.session_bahaya_count = 0
        self.anomaly_events = []

        self.recording = False
        self.csv_file = None
        self.csv_writer = None
        self.ser = None
        self.serial_connected = False

        self.machines = [
            {"name": "Blower Industri UMKM", "icon": "🌀", "bearing_cmd": "B", "rpm_cluster": "Medium (1000-3000 RPM)", "baseline_d2": 9.49, "fw_cluster": "Klaster B"},
            {"name": "Motor Induksi Pompa Air", "icon": "💧", "bearing_cmd": "A", "rpm_cluster": "Low (< 1000 RPM)", "baseline_d2": 8.00, "fw_cluster": "Klaster A"},
            {"name": "Kompresor Production", "icon": "🗜️", "bearing_cmd": "B", "rpm_cluster": "High (> 3000 RPM)", "baseline_d2": 11.00, "fw_cluster": "Klaster B"},
            {"name": "Mesin Blender", "icon": "🥤", "bearing_cmd": "N", "rpm_cluster": "High (> 3000 RPM)", "baseline_d2": 12.50, "fw_cluster": "Klaster B"},
        ]
        self.selected_machine_idx = -1
        self.machine_delete_mode = False
        self.is_technician_mode = True  # Default awal Mode Engineer/Teknisi

        root = QVBoxLayout(self)
        root.setContentsMargins(2, 2, 2, 2)
        root.setSpacing(2)

        # =========================================================================
        # HEADER UTAMA (DENGAN TOMBOL TOGGLE MODE AWAM <-> ENGINEER)
        # =========================================================================
        header = QHBoxLayout()
        header.setSpacing(1)
        header.setContentsMargins(2, 1, 2, 1)
        self.header_frame = QFrame()
        self.header_frame.setStyleSheet(
            f"background-color: {COL_HEADER_BG}; border-radius: 3px; "
            f"border-bottom: 2px solid {COL_ACCENT};"
        )
        self.header_frame.setLayout(header)

        self.lbl_machine_active = QPushButton("⚙️  - Pilih Mesin -")
        self.lbl_machine_active.setStyleSheet(
            f"QPushButton {{ background-color: {COL_ACCENT_DIM}; color: {COL_TEXT_LIGHT}; font-size: 7px; "
            f"font-weight: bold; padding: 1px 2px; border: 1px solid #2a3542; border-radius: 2px; text-align: left; }}"
            f"QPushButton:hover {{ border: 1px solid {COL_ACCENT}; }}"
        )
        self.lbl_machine_active.clicked.connect(lambda: self._change_page(4))
        header.addWidget(self.lbl_machine_active, 2)

        def _header_btn_style(border_color, text_color):
            return (
                f"QPushButton {{ background-color: transparent; color: {text_color}; "
                f"font-weight: bold; font-size: 6px; padding: 1px 2px; "
                f"border: 1px solid {border_color}; border-radius: 2px; }}"
                f"QPushButton:hover {{ background-color: {border_color}; color: #101216; }}"
            )

        self.btn_cek1m = QPushButton("SESI(K)")
        self.btn_cek1m.setStyleSheet(_header_btn_style(COL_OK, COL_OK))
        self.btn_cek1m.clicked.connect(lambda: self._send_command('K'))
        header.addWidget(self.btn_cek1m)

        self.btn_slot_res = QPushButton("SLOT(P)")
        self.btn_slot_res.setStyleSheet(_header_btn_style(COL_ACCENT, COL_ACCENT))
        self.btn_slot_res.clicked.connect(lambda: self._send_command('P'))
        header.addWidget(self.btn_slot_res)

        self.btn_del_base = QPushButton("BASE(Z)")
        self.btn_del_base.setStyleSheet(_header_btn_style(COL_WARN, COL_WARN))
        self.btn_del_base.clicked.connect(lambda: self._send_command('Z'))
        header.addWidget(self.btn_del_base)

        self.btn_recal = QPushButton("KALIBRASI")
        self.btn_recal.setStyleSheet(_header_btn_style(COL_ACCENT, COL_ACCENT))
        self.btn_recal.clicked.connect(lambda: self._send_command('R'))
        header.addWidget(self.btn_recal)

        self.btn_reboot_esp = QPushButton("REBOOT")
        self.btn_reboot_esp.setStyleSheet(_header_btn_style(COL_BAD, COL_BAD))
        self.btn_reboot_esp.clicked.connect(lambda: self._send_command('X'))
        header.addWidget(self.btn_reboot_esp)

        self.btn_debug = QPushButton("DEBUG")
        self.btn_debug.setStyleSheet(_header_btn_style(COL_TEXT_DIM, COL_TEXT_DIM))
        self.btn_debug.clicked.connect(self._show_debug_info)
        header.addWidget(self.btn_debug)

        # Tombol Toggle Mode Dinamis (Awam vs Engineer)
        self.btn_mode_toggle = QPushButton("MODE: ENGINEER")
        self.btn_mode_toggle.setStyleSheet(
            "QPushButton { background-color: #fbbf24; color: #101216; font-weight: bold; font-size: 6px; padding: 2px; border-radius: 2px; border: 1px solid #fbbf24; }"
            "QPushButton:hover { background-color: #f59e0b; }"
        )
        self.btn_mode_toggle.clicked.connect(self._toggle_mode_system)
        header.addWidget(self.btn_mode_toggle)

        self.lbl_conn_dot = QLabel("●")
        self.lbl_conn_dot.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_BAD};")
        header.addWidget(self.lbl_conn_dot)

        time_box = QVBoxLayout()
        time_box.setSpacing(0)
        time_box.setContentsMargins(0, 0, 0, 0)

        self.time_lbl = QLabel("--:--:--")
        self.time_lbl.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_OK};")
        self.time_lbl.setAlignment(Qt.AlignRight)
        time_box.addWidget(self.time_lbl)

        self.date_lbl = QLabel("--/--/----")
        self.date_lbl.setStyleSheet(f"font-size: 5px; color: {COL_OK};")
        self.date_lbl.setAlignment(Qt.AlignRight)
        time_box.addWidget(self.date_lbl)

        header.addLayout(time_box)
        root.addWidget(self.header_frame)

        # =========================================================================
        # STACKED WIDGET (PENGGABUNGAN SEMUA HALAMAN AWAM & ENGINEER)
        # =========================================================================
        self.stack = QStackedWidget()
        
        # Stack Index 0-3: Halaman Mode Awam (Beranda, Riwayat, Rekomendasi, Mesin Saya)
        self.stack.addWidget(self._page_awam_beranda())     # Index 0
        self.stack.addWidget(self._page_awam_riwayat())     # Index 1
        self.stack.addWidget(self._page_awam_rekomendasi()) # Index 2
        self.stack.addWidget(self._page_awam_mesin_saya())  # Index 3
        
        # Stack Index 4-9: Halaman Mode Engineer / Teknisi Asli (Pilih Mesin, Log Detail, Raw Reading, Logs/Recording, Processed FFT, Summary)
        self.stack.addWidget(self._page_machine_select())   # Index 4
        self.stack.addWidget(self._page_log_detail())       # Index 5
        self.stack.addWidget(self._page_raw())              # Index 6
        self.stack.addWidget(self._page_recording())        # Index 7
        self.stack.addWidget(self._page_processed())        # Index 8
        self.stack.addWidget(self._page_summary())          # Index 9
        
        root.addWidget(self.stack, 1)

        # Navigasi Menu Bawah (Dinamis Berubah Sesuai Mode)
        nav_bottom = QHBoxLayout()
        nav_bottom.setSpacing(2)
        
        self.btn_nav1 = QPushButton("RAW READING")
        self.btn_nav2 = QPushButton("LOGS SAVES")
        self.btn_nav3 = QPushButton("PROCESSED (FFT)")
        self.btn_nav4 = QPushButton("SUMMARY")

        self.menu_buttons = [self.btn_nav1, self.btn_nav2, self.btn_nav3, self.btn_nav4]

        for i, btn in enumerate(self.menu_buttons):
            btn.setFixedHeight(30)  
            btn.setStyleSheet(self._menu_style(False))
            btn.clicked.connect(lambda checked, idx=i: self._change_page_by_nav(idx))
            nav_bottom.addWidget(btn)
        
        root.addLayout(nav_bottom)

        self._init_serial_connection()

        self.main_timer = QTimer(self)
        self.main_timer.timeout.connect(self._update_gui)
        self.main_timer.start(200) 

        # Inisialisasi Tampilan Awal Mode Engineer
        self._apply_mode_ui_layout()
        self._refresh_log_list()
        self.show()

    def _toggle_mode_system(self):
        self.is_technician_mode = not self.is_technician_mode
        self._apply_mode_ui_layout()

    def _apply_mode_ui_layout(self):
        if self.is_technician_mode:
            self.btn_mode_toggle.setText("MODE: ENGINEER")
            self.btn_mode_toggle.setStyleSheet(
                "QPushButton { background-color: #fbbf24; color: #101216; font-weight: bold; font-size: 6px; padding: 2px; border-radius: 2px; border: 1px solid #fbbf24; }"
                "QPushButton:hover { background-color: #f59e0b; }"
            )
            self.btn_nav1.setText("RAW READING")
            self.btn_nav2.setText("LOGS SAVES")
            self.btn_nav3.setText("PROCESSED (FFT)")
            self.btn_nav4.setText("SUMMARY")
            self._change_page(6)  # Langsung ke Raw Reading Engineer
        else:
            self.btn_mode_toggle.setText("MODE: AWAM")
            self.btn_mode_toggle.setStyleSheet(
                "QPushButton { background-color: #38bdf8; color: #101216; font-weight: bold; font-size: 6px; padding: 2px; border-radius: 2px; border: 1px solid #38bdf8; }"
                "QPushButton:hover { background-color: #0ea5e9; }"
            )
            self.btn_nav1.setText("BERANDA")
            self.btn_nav2.setText("RIWAYAT")
            self.btn_nav3.setText("REKOMENDASI")
            self.btn_nav4.setText("MESIN SAYA")
            self._change_page(0)  # Langsung ke Beranda Awam

    def _menu_style(self, active):
        if active:
            return f"background-color: {COL_ACCENT}; color: #000000; font-size: 9px; font-weight: bold; border: 1px solid white; border-radius: 3px;"
        return f"background-color: #cfcfcf; color: #000000; font-size: 9px; font-weight: bold; border: 1px solid #444; border-radius: 3px;"

    def _change_page(self, idx):
        self.stack.setCurrentIndex(idx)

    def _change_page_by_nav(self, nav_idx):
        if self.is_technician_mode:
            # Mapping Tombol Navigasi Bawah Mode Engineer (Raw, Logs, FFT, Summary)
            stack_mapping = {0: 6, 1: 7, 2: 8, 3: 9}
            target_stack = stack_mapping.get(nav_idx, 6)
            self._change_page(target_stack)
            for i, btn in enumerate(self.menu_buttons):
                btn.setStyleSheet(self._menu_style(i == nav_idx))
        else:
            # Mapping Tombol Navigasi Bawah Mode Awam (Beranda, Riwayat, Rekomendasi, Mesin Saya)
            target_stack = nav_idx
            self._change_page(target_stack)
            for i, btn in enumerate(self.menu_buttons):
                btn.setStyleSheet(self._menu_style(i == nav_idx))

    # =========================================================================
    # HALAMAN MODE AWAM
    # =========================================================================
    def _page_awam_beranda(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(6)

        self.lbl_beranda_sub = QLabel("Mesin: - Pilih Slot -")
        self.lbl_beranda_sub.setAlignment(Qt.AlignCenter)
        self.lbl_beranda_sub.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_TEXT_DIM};")
        lay.addWidget(self.lbl_beranda_sub)

        box_status = QFrame()
        box_status.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        b_lay = QVBoxLayout(box_status)
        b_lay.setAlignment(Qt.AlignCenter)
        
        self.lbl_beranda_icon = QLabel("?")
        self.lbl_beranda_icon.setAlignment(Qt.AlignCenter)
        self.lbl_beranda_icon.setStyleSheet("font-size: 26px; color: #a0aec0; border: none;")
        b_lay.addWidget(self.lbl_beranda_icon)
        lay.addWidget(box_status, 2)

        self.box_check_status = QFrame()
        self.box_check_status.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        c_lay = QVBoxLayout(self.box_check_status)
        c_lay.setAlignment(Qt.AlignCenter)
        
        self.lbl_check_text = QLabel("BELUM ADA HASIL CHECK")
        self.lbl_check_text.setAlignment(Qt.AlignCenter)
        self.lbl_check_text.setStyleSheet("font-size: 9px; font-weight: bold; color: #e2e8f0; border: none;")
        c_lay.addWidget(self.lbl_check_text)
        lay.addWidget(self.box_check_status, 1)

        box_desc = QFrame()
        box_desc.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        d_lay = QVBoxLayout(box_desc)
        d_lay.setAlignment(Qt.AlignCenter)
        
        self.lbl_desc_text = QLabel("Pilih slot mesin, lalu jalankan sesi Check untuk melihat hasilnya di sini.")
        self.lbl_desc_text.setAlignment(Qt.AlignCenter)
        self.lbl_desc_text.setWordWrap(True)
        self.lbl_desc_text.setStyleSheet(f"font-size: 7px; color: {COL_TEXT_DIM}; border: none;")
        d_lay.addWidget(self.lbl_desc_text)
        lay.addWidget(box_desc, 1)

        self.lbl_estimasi_servis = QLabel("Estimasi servis berikutnya: --")
        self.lbl_estimasi_servis.setAlignment(Qt.AlignCenter)
        self.lbl_estimasi_servis.setStyleSheet(f"font-size: 7px; font-weight: bold; color: {COL_TEXT_LIGHT}; background-color: {COL_PANEL_DARK}; padding: 4px; border-radius: 4px; border: 1px solid #2a3542;")
        lay.addWidget(self.lbl_estimasi_servis)

        return page

    def _page_awam_riwayat(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(6)

        lbl_title = QLabel("Riwayat Sesi Pemantauan Ini")
        lbl_title.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        lay.addWidget(lbl_title)

        grid_stat = QGridLayout()
        grid_stat.setSpacing(4)

        def _make_stat_box(title):
            f = QFrame()
            f.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 4px;")
            flay = QVBoxLayout(f)
            flay.setContentsMargins(2, 4, 2, 4)
            flay.setAlignment(Qt.AlignCenter)
            
            t = QLabel(title)
            t.setStyleSheet("font-size: 6px; color: #94a3b8; border: none;")
            t.setAlignment(Qt.AlignCenter)
            
            v = QLabel("0")
            v.setStyleSheet("font-size: 11px; font-weight: bold; color: #ffffff; border: none;")
            v.setAlignment(Qt.AlignCenter)
            
            flay.addWidget(t)
            flay.addWidget(v)
            return f, v

        box_w, self.lbl_stat_waspada = _make_stat_box("Kali Waspada")
        box_b, self.lbl_stat_bahaya = _make_stat_box("Kali Bahaya")
        box_t, self.lbl_stat_total = _make_stat_box("Total Data Masuk")

        grid_stat.addWidget(box_w, 0, 0)
        grid_stat.addWidget(box_b, 0, 1)
        grid_stat.addWidget(box_t, 0, 2)
        lay.addLayout(grid_stat)

        lbl_kejadian = QLabel("Kejadian Terakhir:")
        lbl_kejadian.setStyleSheet(f"font-size: 7px; color: {COL_TEXT_DIM};")
        lay.addWidget(lbl_kejadian)

        self.list_riwayat = QListWidget()
        self.list_riwayat.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: #ffffff; border: 1px solid #2a3542; border-radius: 4px; font-size: 7px;")
        self.list_riwayat.addItem("Belum ada kejadian pada sesi ini.")
        lay.addWidget(self.list_riwayat, 1)

        return page

    def _page_awam_rekomendasi(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(6)

        lbl_title = QLabel("Rekomendasi Tindakan")
        lbl_title.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        lay.addWidget(lbl_title)

        box_rec = QFrame()
        box_rec.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        r_lay = QVBoxLayout(box_rec)
        r_lay.setContentsMargins(8, 8, 8, 8)
        r_lay.setSpacing(4)

        self.lbl_rec_penyebab = QLabel("Kemungkinan penyebab: belum ada hasil Check")
        self.lbl_rec_penyebab.setStyleSheet(f"font-size: 7px; font-weight: bold; color: {COL_WARN}; border: none;")
        self.lbl_rec_penyebab.setWordWrap(True)
        r_lay.addWidget(self.lbl_rec_penyebab)

        self.lbl_rec_tindakan = QLabel("Tindakan yang disarankan: pilih slot mesin lalu jalankan sesi Check.")
        self.lbl_rec_tindakan.setStyleSheet(f"font-size: 7px; color: {COL_TEXT_LIGHT}; border: none;")
        self.lbl_rec_tindakan.setWordWrap(True)
        r_lay.addWidget(self.lbl_rec_tindakan)
        r_lay.addStretch(1)

        lay.addWidget(box_rec, 1)
        return page

    def _page_awam_mesin_saya(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(6)

        lbl_title = QLabel("Pilih Slot Mesin")
        lbl_title.setAlignment(Qt.AlignCenter)
        lbl_title.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        lay.addWidget(lbl_title)

        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(6)

        self.card_m1 = self._make_machine_card_widget(0)
        self.card_m2 = self._make_machine_card_widget(1)

        cards_layout.addWidget(self.card_m1)
        cards_layout.addWidget(self.card_m2)
        lay.addLayout(cards_layout, 1)

        return page

    def _make_machine_card_widget(self, idx):
        m = self.machines[idx]
        frame = QFrame()
        frame.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid {COL_OK}; border-radius: 6px;")
        
        flay = QVBoxLayout(frame)
        flay.setAlignment(Qt.AlignCenter)
        flay.setSpacing(4)

        lbl_icon = QLabel(m["icon"])
        lbl_icon.setStyleSheet("font-size: 16px; border: none;")
        lbl_icon.setAlignment(Qt.AlignCenter)
        flay.addWidget(lbl_icon)

        lbl_name = QLabel(m["name"])
        lbl_name.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_TEXT_LIGHT}; border: none;")
        lbl_name.setAlignment(Qt.AlignCenter)
        flay.addWidget(lbl_name)

        btn_stat = QPushButton("PILIH MESIN")
        btn_stat.setStyleSheet(f"background-color: {COL_ACCENT}; color: #101216; font-size: 6px; font-weight: bold; border-radius: 3px; padding: 3px;")
        btn_stat.clicked.connect(lambda checked, i=idx: self._select_machine_slot(i))
        flay.addWidget(btn_stat)

        return frame

    def _select_machine_slot(self, idx):
        self.selected_machine_idx = idx
        m = self.machines[idx]
        self.lbl_machine_active.setText(f"{m['icon']}  {m['name']}")
        self.lbl_beranda_sub.setText(f"Mesin: {m['name']}")
        self._send_command(m["bearing_cmd"])
        if not self.is_technician_mode:
            self._change_page(0)

    # =========================================================================
    # HALAMAN MODE ENGINEER / TEKNISI (RAW, RECORDING, FFT, SUMMARY, DLL)
    # =========================================================================
    def _page_raw(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(1, 1, 1, 1)
        layout.setSpacing(1)

        main_raw = QHBoxLayout()
        main_raw.setSpacing(2)
        
        layout_grafik_grid = QGridLayout()
        layout_grafik_grid.setSpacing(1)
        layout_grafik_grid.setContentsMargins(0, 0, 0, 0)
        
        pg.setConfigOptions(antialias=True)

        def _tidy_plot(plot_widget, title):
            axis_font = QFont("Arial", 7)
            plot_item = plot_widget.getPlotItem()
            for axis_name in ("left", "bottom"):
                axis = plot_item.getAxis(axis_name)
                axis.setStyle(tickFont=axis_font, tickTextOffset=3, autoExpandTextSpace=True)
                axis.setTextPen(pg.mkPen('#dddddd'))
            plot_item.getAxis("bottom").setTickSpacing(major=15, minor=15)
            plot_item.getAxis("left").setWidth(30)
            plot_widget.setTitle(title, size="8pt")
            plot_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.graph_a = pg.PlotWidget(title="Sound (dB)")
        self.graph_a.setBackground(COL_PANEL_DARK)
        self.graph_a.showGrid(x=True, y=True, alpha=0.2)
        self.curve_a = self.graph_a.plot(pen=pg.mkPen('#a78bfa', width=1.5))
        _tidy_plot(self.graph_a, "Sound (dB)")
        layout_grafik_grid.addWidget(self.graph_a, 0, 0)
        
        self.graph_temp = pg.PlotWidget(title="Temp (°C)")
        self.graph_temp.setBackground(COL_PANEL_DARK)
        self.graph_temp.showGrid(x=True, y=True, alpha=0.2)
        self.curve_temp = self.graph_temp.plot(pen=pg.mkPen('#f472b6', width=1.5))
        _tidy_plot(self.graph_temp, "Temp (°C)")
        layout_grafik_grid.addWidget(self.graph_temp, 0, 1)

        self.graph_v = pg.PlotWidget(title="Vibration (G)")
        self.graph_v.setBackground(COL_PANEL_DARK)
        self.graph_v.showGrid(x=True, y=True, alpha=0.2)
        self.curve_v = self.graph_v.plot(pen=pg.mkPen('#818cf8', width=1.5))
        _tidy_plot(self.graph_v, "Vibration (G)")
        layout_grafik_grid.addWidget(self.graph_v, 1, 0)
        
        self.panel_ai = QFrame()
        self.panel_ai.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border-radius: 3px;")
        lay_ai = QVBoxLayout(self.panel_ai)
        lay_ai.setContentsMargins(3, 1, 3, 1)
        lay_ai.setSpacing(1)

        lbl_ai_title = QLabel("🤖 ADVANCED DIAGNOSTICS & AI")
        lbl_ai_title.setStyleSheet(f"font-size: 7px; font-weight: bold; color: {COL_ACCENT};")
        lay_ai.addWidget(lbl_ai_title)

        grid_ai = QGridLayout()
        grid_ai.setHorizontalSpacing(3)
        grid_ai.setVerticalSpacing(0)

        def _ai_row(r, label_text):
            lbl1 = QLabel(label_text)
            lbl1.setStyleSheet("font-size: 6px; color: #999;")
            lbl2 = QLabel("--")
            lbl2.setStyleSheet("font-size: 7px; font-weight: bold; color: #eee;")
            lbl2.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            grid_ai.addWidget(lbl1, r, 0)
            grid_ai.addWidget(lbl2, r, 1)
            return lbl2

        self.lbl_hs_val = _ai_row(0, "Health Score:")
        self.lbl_ml_val = _ai_row(1, "TinyML Mode:")
        self.lbl_diag_val = _ai_row(2, "Fault Diag:")
        self.lbl_kurt_val = _ai_row(3, "Kurtosis:")
        self.lbl_flags_val = _ai_row(4, "Multi-Fault:")

        self.prog_hs = QProgressBar()
        self.prog_hs.setFixedHeight(3)
        self.prog_hs.setTextVisible(False)
        self.prog_hs.setStyleSheet(
            "QProgressBar { border: none; background-color: #333; border-radius: 1px; }"
            "QProgressBar::chunk { background-color: " + COL_OK + "; border-radius: 1px; }"
        )
        grid_ai.addWidget(self.prog_hs, 5, 0, 1, 2)

        lay_ai.addLayout(grid_ai)
        lay_ai.addStretch(1)
        layout_grafik_grid.addWidget(self.panel_ai, 1, 1)

        layout_grafik_grid.setColumnStretch(0, 1)
        layout_grafik_grid.setColumnStretch(1, 1)
        layout_grafik_grid.setRowStretch(0, 1)
        layout_grafik_grid.setRowStretch(1, 1)
        
        main_raw.addLayout(layout_grafik_grid, 5)

        panel_kanan = QVBoxLayout()
        panel_kanan.setSpacing(1)

        frame_status = QFrame()
        frame_status.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border-radius: 3px;")
        fs_lay = QVBoxLayout(frame_status)
        fs_lay.setContentsMargins(3, 2, 3, 2)
        fs_lay.setSpacing(1)

        self.lbl_sys_status = QLabel("● STANDBY")
        self.lbl_sys_status.setStyleSheet("font-size: 7px; font-weight: bold; color: #888888; padding-bottom: 1px;")
        fs_lay.addWidget(self.lbl_sys_status)

        row_style = "border-bottom: 1px solid #33363b;"
        name_style = "font-size: 7px; color: #999999;"
        val_style = "font-size: 9px; color: #eeeeee; font-weight: bold;"

        grid_val = QGridLayout()
        grid_val.setContentsMargins(0, 0, 0, 0)
        grid_val.setHorizontalSpacing(3)
        grid_val.setVerticalSpacing(1)
        grid_val.setColumnStretch(0, 1)
        grid_val.setColumnStretch(1, 1)

        def _param_row(row, name):
            lbl_name = QLabel(name)
            lbl_name.setStyleSheet(name_style + row_style + "padding: 1px 0px;")
            lbl_val = QLabel("--")
            lbl_val.setStyleSheet(val_style + row_style + "padding: 1px 0px;")
            lbl_val.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            grid_val.addWidget(lbl_name, row, 0)
            grid_val.addWidget(lbl_val, row, 1)
            return lbl_val

        self.lbl_val_v = _param_row(0, "Vibration")
        self.lbl_val_a = _param_row(1, "Sound")
        self.lbl_val_temp = _param_row(2, "Temp")
        self.lbl_val_rpm = _param_row(3, "RPM")
        self.lbl_val_d2 = _param_row(4, "D² (Severity)")

        fs_lay.addLayout(grid_val)

        self.lbl_val_vxyz = QLabel("(X: -- | Y: -- | Z: -- G)")
        self.lbl_val_vxyz.setStyleSheet("font-size: 6px; color: #777777; padding: 1px 0px 0px 0px;")
        self.lbl_val_vxyz.setWordWrap(True)
        fs_lay.addWidget(self.lbl_val_vxyz)

        panel_kanan.addWidget(frame_status)
        main_raw.addLayout(panel_kanan, 4)

        layout.addLayout(main_raw)
        return page

    def _page_recording(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        rec_ctrl = QHBoxLayout()
        self.btn_toggle_rec = QPushButton("MULAI RECORDING")
        self.btn_toggle_rec.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 7px; height: 20px; border-radius: 3px;")
        self.btn_toggle_rec.clicked.connect(self._toggle_recording)
        
        self.lbl_rec_status = QLabel("● Device Standby")
        self.lbl_rec_status.setStyleSheet("font-size: 6px; color: #aaa;")
        
        rec_ctrl.addWidget(self.btn_toggle_rec, 2)
        rec_ctrl.addWidget(self.lbl_rec_status, 1)
        layout.addLayout(rec_ctrl)

        lbl_daftar = QLabel("Daftar Rekaman Mesin (.CSV):")
        lbl_daftar.setStyleSheet("font-size: 6px;")
        layout.addWidget(lbl_daftar)
        
        self.log_list = QListWidget()
        self.log_list.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: white; font-size: 7px;")
        self.log_list.itemDoubleClicked.connect(self._open_log_detail)
        layout.addWidget(self.log_list)

        btn_lay = QHBoxLayout()
        self.btn_watch_log = QPushButton("Buka Panel Deteksi")
        self.btn_export_excel = QPushButton("Export ke Excel")
        self.btn_delete_log = QPushButton("Hapus Rekaman")

        self.btn_watch_log.setStyleSheet("background-color: #cfcfcf; color: #000000; font-size: 7px; height: 18px; font-weight: bold; border-radius: 2px;")
        self.btn_export_excel.setStyleSheet("background-color: #217346; color: #ffffff; font-size: 7px; height: 18px; font-weight: bold; border-radius: 2px;")
        self.btn_delete_log.setStyleSheet("background-color: #cfcfcf; color: #000000; font-size: 7px; height: 18px; font-weight: bold; border-radius: 2px;")

        self.btn_watch_log.clicked.connect(self._open_log_detail_from_button)
        self.btn_export_excel.clicked.connect(self._export_selected_log_to_excel)
        self.btn_delete_log.clicked.connect(self._delete_selected_log)

        btn_lay.addWidget(self.btn_watch_log)
        btn_lay.addWidget(self.btn_export_excel)
        btn_lay.addWidget(self.btn_delete_log)
        layout.addLayout(btn_lay)

        lbl_hint = QLabel("* Klik file rekaman untuk membuka panel forensik anomali.")
        lbl_hint.setStyleSheet("font-size: 6px; color: #888; font-style: italic;")
        layout.addWidget(lbl_hint)

        return page

    def _page_processed(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        self.lbl_proc_snapshot = QLabel("Menunggu data... | RPM: -- | D²: --")
        self.lbl_proc_snapshot.setStyleSheet("font-size: 6px; color: #999;")
        layout.addWidget(self.lbl_proc_snapshot)

        grid = QGridLayout()
        grid.setSpacing(2)
        pg.setConfigOptions(antialias=True)

        self.graph_fft = pg.PlotWidget(title="Frequency Spectrum (FFT)")
        self.graph_fft.setBackground(COL_PANEL_DARK)
        self.graph_fft.showGrid(x=True, y=True, alpha=0.3)
        self.curve_fft = self.graph_fft.plot(pen=pg.mkPen('#a855f7', width=1.2))
        grid.addWidget(self.graph_fft, 0, 0, 1, 2)

        self.graph_rpm = pg.PlotWidget(title="RPM Estimasi")
        self.graph_rpm.setBackground(COL_PANEL_DARK)
        self.graph_rpm.showGrid(x=True, y=True, alpha=0.3)
        self.curve_rpm = self.graph_rpm.plot(pen=pg.mkPen('#4da6ff', width=1.2))
        grid.addWidget(self.graph_rpm, 1, 0)

        self.graph_d2 = pg.PlotWidget(title="Mahalanobis D²")
        self.graph_d2.setBackground(COL_PANEL_DARK)
        self.graph_d2.showGrid(x=True, y=True, alpha=0.3)
        self.curve_d2 = self.graph_d2.plot(pen=pg.mkPen('#ff6666', width=1.2))
        
        line_waspada = pg.InfiniteLine(pos=D2_THRESHOLD_WASPADA, angle=0, pen=pg.mkPen(COL_WARN, width=1, style=Qt.DashLine))
        line_bahaya = pg.InfiniteLine(pos=D2_THRESHOLD_BAHAYA, angle=0, pen=pg.mkPen(COL_BAD, width=1, style=Qt.DashLine))
        self.graph_d2.addItem(line_waspada)
        self.graph_d2.addItem(line_bahaya)
        grid.addWidget(self.graph_d2, 1, 1)

        layout.addLayout(grid, 3)

        lbl_anomali_title = QLabel("Log Kejadian Anomali:")
        lbl_anomali_title.setStyleSheet("font-size: 6px; font-weight: bold; color: #ccc;")
        layout.addWidget(lbl_anomali_title)

        self.list_anomali = QListWidget()
        self.list_anomali.setStyleSheet("background-color: #ffffff; color: #222222; font-size: 7px;")
        self.list_anomali.addItem("Tidak ada kejadian anomali sepanjang sesi ini.")
        layout.addWidget(self.list_anomali, 2)

        self.lbl_session_summary = QLabel(
            "Sesi: 0 sample | RPM rata-rata: 0.0 | D² max: 0.00 | Kondisi terparah: Normal | Waspada: 0x, Bahaya: 0x."
        )
        self.lbl_session_summary.setStyleSheet(
            "font-size: 7px; color: #1c3d1c; background-color: #d7f0d7; padding: 2px; border-radius: 2px;"
        )
        self.lbl_session_summary.setWordWrap(True)
        layout.addWidget(self.lbl_session_summary)

        return page

    def _page_summary(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)

        self.lbl_sum_machine = QLabel("Target: Belum dipilih")
        self.lbl_sum_machine.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        layout.addWidget(self.lbl_sum_machine)

        grid = QGridLayout()
        grid.setSpacing(4)

        self.gauge_s = GradientGauge("Sound", 30.0, 100.0, 75.0, 85.0, "dB")
        self.gauge_t = GradientGauge("Temperature", 20.0, 80.0, 42.0, 50.0, "°C")
        self.gauge_v = GradientGauge("Vibration", 0.0, 0.5, 0.18, 0.25, "G")

        grid.addWidget(self.gauge_s, 0, 0)
        grid.addWidget(self.gauge_t, 0, 1)
        grid.addWidget(self.gauge_v, 1, 0)

        self.reserved_slot_sum = QFrame()
        self.reserved_slot_sum.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px dashed {COL_TEXT_DIM}; border-radius: 3px;")
        grid.addWidget(self.reserved_slot_sum, 1, 1)

        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        grid.setRowStretch(0, 1)
        grid.setRowStretch(1, 1)

        layout.addLayout(grid)
        layout.addStretch(1)

        self.lbl_diag_desc_summary = QLabel("Menunggu data deteksi...")
        self.lbl_diag_desc_summary.setStyleSheet(f"font-size: 8px; color: {COL_TEXT_DIM}; border-top: 1px solid #333; padding-top: 2px;")
        self.lbl_diag_desc_summary.setWordWrap(True)
        self.lbl_diag_desc_summary.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.lbl_diag_desc_summary)

        return page

    def _page_machine_select(self):
        page = QWidget()
        self.machine_page_layout = QVBoxLayout(page)
        self.machine_page_layout.setContentsMargins(4, 4, 4, 4)
        self.machine_page_layout.setSpacing(2)

        top_row = QHBoxLayout()
        lbl_title = QLabel("Pilih Mesin Target & Klaster Firmware (Klaster A/B)")
        lbl_title.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        top_row.addWidget(lbl_title)
        top_row.addStretch(1)

        self.btn_machine_delete_mode = QPushButton("🗑 HAPUS")
        self.btn_machine_delete_mode.setCheckable(True)
        self.btn_machine_delete_mode.setStyleSheet(self._delete_mode_btn_style(False))
        self.btn_machine_delete_mode.clicked.connect(self._toggle_machine_delete_mode)
        top_row.addWidget(self.btn_machine_delete_mode)
        self.machine_page_layout.addLayout(top_row)

        self.machine_grid = QGridLayout()
        self.machine_grid.setSpacing(3)
        for c in range(3):
            self.machine_grid.setColumnStretch(c, 1)
        for r in range(2):
            self.machine_grid.setRowStretch(r, 1)
        self.machine_page_layout.addLayout(self.machine_grid, 1)

        self._rebuild_machine_grid()
        return page

    def _delete_mode_btn_style(self, active):
        if active:
            return f"background-color: {COL_BAD}; color: #101216; font-weight: bold; font-size: 7px; padding: 2px; border-radius: 2px;"
        return f"background-color: transparent; color: {COL_BAD}; font-weight: bold; font-size: 7px; padding: 2px; border: 1px solid {COL_BAD}; border-radius: 2px;"

    def _toggle_machine_delete_mode(self):
        self.machine_delete_mode = self.btn_machine_delete_mode.isChecked()
        self.btn_machine_delete_mode.setStyleSheet(self._delete_mode_btn_style(self.machine_delete_mode))

    def _make_machine_card(self, idx):
        btn = QPushButton()
        btn.setFixedHeight(55)
        lay = QVBoxLayout(btn)
        lay.setContentsMargins(1, 1, 1, 1)
        lay.setSpacing(0)

        if idx is None:
            lbl_icon = QLabel("➕")
            lbl_name = QLabel("Tambah Mesin")
            btn.clicked.connect(self._activate_inline_add_machine)
            border_col = COL_TEXT_DIM
            bg = "transparent"
            lbl_cluster = QLabel("")
        else:
            m = self.machines[idx]
            lbl_icon = QLabel(m["icon"])
            lbl_name = QLabel(m["name"])
            cluster_text = f"{m.get('rpm_cluster', '')} ({m.get('fw_cluster', 'Klaster B')})"
            lbl_cluster = QLabel(f"⚡ {cluster_text}")
            btn.clicked.connect(lambda checked, i=idx: self._on_machine_card_clicked(i))
            selected = (idx == self.selected_machine_idx)
            border_col = COL_ACCENT if selected else "#2a3542"
            bg = COL_ACCENT_DIM if selected else COL_PANEL_DARK

        lbl_icon.setAlignment(Qt.AlignCenter)
        lbl_icon.setStyleSheet("font-size: 14px; background: transparent; border: none;")
        
        lbl_name.setAlignment(Qt.AlignCenter)
        lbl_name.setWordWrap(True)
        lbl_name.setStyleSheet(f"font-size: 6px; font-weight: bold; color: {COL_TEXT_LIGHT}; background: transparent; border: none;")

        lbl_cluster.setAlignment(Qt.AlignCenter)
        lbl_cluster.setStyleSheet(f"font-size: 5px; color: {COL_ACCENT}; background: transparent; border: none;")

        lay.addWidget(lbl_icon)
        lay.addWidget(lbl_name)
        lay.addWidget(lbl_cluster)

        btn.setStyleSheet(
            f"QPushButton {{ background-color: {bg}; border: 1px solid {border_col}; border-radius: 3px; }}"
            f"QPushButton:hover {{ border: 1px solid {COL_ACCENT}; }}"
        )
        return btn

    def _rebuild_machine_grid(self):
        while self.machine_grid.count():
            item = self.machine_grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        total_slots = 6
        for i in range(min(len(self.machines), total_slots)):
            self.machine_grid.addWidget(self._make_machine_card(i), i // 3, i % 3)

        if len(self.machines) < total_slots:
            add_pos = len(self.machines)
            self.machine_grid.addWidget(self._make_machine_card(None), add_pos // 3, add_pos % 3)

    def _activate_inline_add_machine(self):
        while self.machine_grid.count():
            item = self.machine_grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        form_frame = QFrame()
        form_frame.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid {COL_ACCENT}; border-radius: 3px;")
        f_lay = QVBoxLayout(form_frame)
        f_lay.setContentsMargins(3, 3, 3, 3)
        f_lay.setSpacing(1)

        lbl_nm = QLabel("Nama Mesin Baru:")
        lbl_nm.setStyleSheet(f"color: {COL_TEXT_LIGHT}; font-size: 6px; background: transparent;")
        f_lay.addWidget(lbl_nm)

        self.inline_name_input = QLineEdit()
        self.inline_name_input.setStyleSheet(f"background-color: {COL_BG_MAIN}; color: {COL_TEXT_LIGHT}; border: 1px solid {COL_TEXT_DIM}; padding: 1px; font-size: 6px;")
        f_lay.addWidget(self.inline_name_input)

        h_sub = QHBoxLayout()
        self.inline_icon_combo = QComboBox()
        self.inline_icon_combo.addItems(["⚙️", "🌀", "💧", "🗜️", "🥤", "🔥", "❄️", "🏭"])
        self.inline_icon_combo.setStyleSheet(f"background-color: {COL_BG_MAIN}; color: {COL_TEXT_LIGHT}; font-size: 6px;")
        h_sub.addWidget(self.inline_icon_combo)

        self.inline_fw_cluster_combo = QComboBox()
        self.inline_fw_cluster_combo.addItems(["Klaster A (≈1400 RPM)", "Klaster B (≈2800 RPM)"])
        self.inline_fw_cluster_combo.setStyleSheet(f"background-color: {COL_BG_MAIN}; color: {COL_TEXT_LIGHT}; font-size: 6px;")
        h_sub.addWidget(self.inline_fw_cluster_combo)
        f_lay.addLayout(h_sub)

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("Simpan")
        btn_cancel = QPushButton("Batal")
        btn_ok.setStyleSheet(f"background-color: {COL_ACCENT}; color: #000; font-weight: bold; font-size: 6px; padding: 2px;")
        btn_cancel.setStyleSheet(f"background-color: #444; color: #fff; font-size: 6px; padding: 2px;")
        
        btn_ok.clicked.connect(self._save_inline_machine)
        btn_cancel.clicked.connect(self._rebuild_machine_grid)
        
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        f_lay.addLayout(btn_row)

        self.machine_grid.addWidget(form_frame, 0, 0, 2, 3)

    def _save_inline_machine(self):
        name = self.inline_name_input.text().strip()
        if not name:
            return
        icon = self.inline_icon_combo.currentText()
        fw_cluster_idx = self.inline_fw_cluster_combo.currentIndex()
        bearing_cmd = "A" if fw_cluster_idx == 0 else "B"
        fw_cluster_name = "Klaster A" if fw_cluster_idx == 0 else "Klaster B"
        rpm_cluster_desc = "Low (< 1000 RPM)" if fw_cluster_idx == 0 else "Medium/High (> 1500 RPM)"

        self.machines.append({
            "name": name, 
            "icon": icon, 
            "bearing_cmd": bearing_cmd, 
            "rpm_cluster": rpm_cluster_desc,
            "baseline_d2": 9.49,
            "fw_cluster": fw_cluster_name
        })
        self._rebuild_machine_grid()

    def _on_machine_card_clicked(self, idx):
        if self.machine_delete_mode:
            m = self.machines[idx]
            self.machines.pop(idx)
            if self.selected_machine_idx == idx:
                self.selected_machine_idx = -1
                self.lbl_machine_active.setText("⚙️  - Pilih Mesin -")
            elif self.selected_machine_idx > idx:
                self.selected_machine_idx -= 1
            self._rebuild_machine_grid()
            return

        self.selected_machine_idx = idx
        m = self.machines[idx]
        self.lbl_machine_active.setText(f"{m['icon']}  {m['name']} ({m.get('fw_cluster', 'Klaster B')})")
        self._send_command(m["bearing_cmd"])
        self._rebuild_machine_grid()

    def _page_log_detail(self):
        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(2, 2, 2, 2)
        root.setSpacing(2)

        self.lbl_logdet_title = QLabel("HASIL DETEKSI REKAMAN: -")
        self.lbl_logdet_title.setStyleSheet("font-size: 8px; font-weight: bold; color: #ffffff;")
        self.lbl_logdet_title.setWordWrap(True)
        root.addWidget(self.lbl_logdet_title)

        status_row = QHBoxLayout()
        status_row.setSpacing(2)

        self.lbl_logdet_dot = QLabel("●")
        self.lbl_logdet_dot.setStyleSheet("font-size: 11px; color: #888888;")
        status_row.addWidget(self.lbl_logdet_dot)

        self.lbl_logdet_peak = QLabel("Nilai puncak: -")
        self.lbl_logdet_peak.setWordWrap(True)
        self.lbl_logdet_peak.setStyleSheet("font-size: 7px; color: #cccccc;")
        status_row.addWidget(self.lbl_logdet_peak, 1)

        root.addLayout(status_row)

        grid = QGridLayout()
        grid.setSpacing(1)
        pg.setConfigOptions(antialias=True)

        self.graph_logdet_a = pg.PlotWidget(title="Sound (dB)")
        self.graph_logdet_a.setBackground(COL_PANEL_DARK)
        self.graph_logdet_a.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_a = self.graph_logdet_a.plot(pen=pg.mkPen('#4da6ff', width=1.2))
        grid.addWidget(self.graph_logdet_a, 0, 0)

        self.graph_logdet_temp = pg.PlotWidget(title="Temp (°C)")
        self.graph_logdet_temp.setBackground(COL_PANEL_DARK)
        self.graph_logdet_temp.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_temp = self.graph_logdet_temp.plot(pen=pg.mkPen('#e040fb', width=1.2))
        grid.addWidget(self.graph_logdet_temp, 0, 1)

        self.graph_logdet_v = pg.PlotWidget(title="Vibration (G)")
        self.graph_logdet_v.setBackground(COL_PANEL_DARK)
        self.graph_logdet_v.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_v = self.graph_logdet_v.plot(pen=pg.mkPen('#ff4d4d', width=1.2))
        grid.addWidget(self.graph_logdet_v, 1, 0)

        self.reserved_slot_logdet = QFrame()
        self.reserved_slot_logdet.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px dashed {COL_TEXT_DIM}; border-radius: 3px;")
        grid.addWidget(self.reserved_slot_logdet, 1, 1)

        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)

        root.addLayout(grid, 20)

        ctrl = QHBoxLayout()
        ctrl.setSpacing(2)

        self.btn_logdet_play_pause = QPushButton("▶ PLAY")
        self.speed_logdet_combo = QComboBox()
        self.speed_logdet_combo.addItems(["0.5x", "1x", "2x", "4x"])
        self.speed_logdet_combo.setCurrentIndex(1)
        self.btn_logdet_back = QPushButton("◄ KEMBALI")

        self.btn_logdet_play_pause.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 7px; height: 20px; border-radius: 3px;")
        self.btn_logdet_back.setStyleSheet(f"background-color: {COL_BAD}; color: #ffffff; font-weight: bold; font-size: 7px; height: 20px; border-radius: 3px;")
        self.speed_logdet_combo.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: white; font-size: 7px;")

        self.btn_logdet_play_pause.clicked.connect(self._logdet_toggle_play)
        self.speed_logdet_combo.currentIndexChanged.connect(self._logdet_change_speed)
        self.btn_logdet_back.clicked.connect(self._logdet_back)

        self.slider_logdet = QSlider(Qt.Horizontal)
        self.slider_logdet.setMinimum(0)
        self.slider_logdet.setMaximum(0)
        self.slider_logdet.sliderMoved.connect(self._logdet_seek)

        self.lbl_logdet_pos = QLabel("0 / 0")
        self.lbl_logdet_pos.setStyleSheet("font-size: 6px; color: #aaa;")

        ctrl.addWidget(self.btn_logdet_play_pause, 2)
        ctrl.addWidget(self.speed_logdet_combo, 1)
        ctrl.addWidget(self.slider_logdet, 5)
        ctrl.addWidget(self.lbl_logdet_pos, 1)
        ctrl.addWidget(self.btn_logdet_back, 2)
        root.addLayout(ctrl)

        return page

    def _logdet_back(self):
        self.logdet_timer.stop()
        self._change_page(7)

    def _logdet_load_csv(self, filepath):
        times, v_vals, a_vals, t_vals = [], [], [], []
        try:
            with open(filepath, 'r') as f:
                reader = csv.reader(f)
                header = next(reader)
                idx_v = header.index('rms_v') if 'rms_v' in header else 2
                idx_a = header.index('rms_a') if 'rms_a' in header else 3
                idx_t = header.index('temp') if 'temp' in header else 4

                for i, row in enumerate(reader):
                    if len(row) <= max(idx_v, idx_a, idx_t): continue
                    times.append(i)
                    v_vals.append(float(row[idx_v]))
                    a_vals.append(float(row[idx_a]))
                    t_vals.append(float(row[idx_t]))
        except Exception as e:
            print(f"Error load csv: {e}")
        return times, v_vals, a_vals, t_vals

    def _logdet_render_diagnosis_summary(self):
        if not self.logdet_v_vals:
            self.lbl_logdet_dot.setStyleSheet("font-size: 11px; color: #888888;")
            self.lbl_logdet_peak.setText("Data kosong.")
            return

        peak_v = max(self.logdet_v_vals)
        peak_a = max(self.logdet_a_vals)
        peak_temp = max(self.logdet_t_vals)

        if peak_v > 0.25 or peak_temp > 50.0:
            status, col = "BAHAYA", COL_BAD
        elif peak_v > 0.18 or peak_temp > 42.0:
            status, col = "WASPADA", COL_WARN
        else:
            status, col = "NORMAL", COL_OK

        self.lbl_logdet_dot.setStyleSheet(f"font-size: 11px; color: {col};")
        self.lbl_logdet_peak.setText(
            f"{status} — Puncak: Vib {peak_v:.2f} G | Snd {peak_a:.1f} dB | Tmp {peak_temp:.1f} °C"
        )

    def _logdet_render_frame(self, idx):
        start = max(0, idx - 49)
        t = self.logdet_times[start:idx + 1]
        v = self.logdet_v_vals[start:idx + 1]
        a = self.logdet_a_vals[start:idx + 1]
        tp = self.logdet_t_vals[start:idx + 1]

        self.curve_logdet_v.setData(t, v)
        self.curve_logdet_a.setData(t, a)
        self.curve_logdet_temp.setData(t, tp)

        self.slider_logdet.blockSignals(True)
        self.slider_logdet.setValue(idx)
        self.slider_logdet.blockSignals(False)
        self.lbl_logdet_pos.setText(f"{idx + 1} / {len(self.logdet_times)}")

    def _logdet_step(self):
        if self.logdet_play_index >= len(self.logdet_times):
            self.logdet_timer.stop()
            self.btn_logdet_play_pause.setText("▶ PLAY")
            return
        self._logdet_render_frame(self.logdet_play_index)
        self.logdet_play_index += 1

    def _logdet_toggle_play(self):
        if not self.logdet_times:
            return
        if self.logdet_timer.isActive():
            self.logdet_timer.stop()
            self.btn_logdet_play_pause.setText("▶ PLAY")
        else:
            if self.logdet_play_index >= len(self.logdet_times):
                self.logdet_play_index = 0
            interval_ms = max(20, int(200 / self.logdet_play_speed))
            self.logdet_timer.start(interval_ms)
            self.btn_logdet_play_pause.setText("⏸ PAUSE")

    def _logdet_change_speed(self, combo_idx):
        mapping = {0: 0.5, 1: 1.0, 2: 2.0, 3: 4.0}
        self.logdet_play_speed = mapping.get(combo_idx, 1.0)
        if self.logdet_timer.isActive():
            self.logdet_timer.start(max(20, int(200 / self.logdet_play_speed)))

    def _logdet_seek(self, value):
        self.logdet_play_index = value
        self._logdet_render_frame(value)
    
    def _init_serial_connection(self):
        if serial is not None:
            t = threading.Thread(target=self._read_serial_worker, daemon=True)
            t.start()

    def _resolve_serial_port(self):
        try:
            ports = list(serial.tools.list_ports.comports())
        except Exception:
            ports = []
        if not ports: 
            return SERIAL_PORT
        available = [p.device for p in ports]
        if SERIAL_PORT in available: 
            return SERIAL_PORT
        for p in ports:
            desc = f"{p.description} {p.manufacturer or ''}".upper()
            if any(hint in desc for hint in ESP32_USB_HINTS):
                return p.device
        return ports[0].device if ports else SERIAL_PORT

    def _read_serial_worker(self):
        while True:
            try:
                if self.ser is None or not self.ser.is_open:
                    port_to_use = self._resolve_serial_port()
                    self.ser = serial.Serial(port_to_use, BAUD_RATE, timeout=0.1)
                    self.serial_connected = True

                raw = self.ser.readline()
                if not raw:
                    if time.time() - self.last_packet_time > 2.5:
                        self.packet_loss_flag = True
                    continue

                self.last_packet_time = time.time()
                self.packet_loss_flag = False

                line = raw.decode('utf-8', errors='ignore').strip()
                if not line.startswith("{"): 
                    continue

                try:
                    data = json.loads(line)
                except (json.JSONDecodeError, ValueError):
                    continue

                self.last_raw_line = line
                self.current_v = float(data.get("rms_v", 0.0))
                self.current_a = float(data.get("rms_a", 0.0))
                self.current_temp = float(data.get("temp", 0.0))
                self.current_vx = float(data.get("vib_x", 0.0))
                self.current_vy = float(data.get("vib_y", 0.0))
                self.current_vz = float(data.get("vib_z", 0.0))
                self.current_rpm = float(data.get("rpm", 0.0))
                self.current_d2 = float(data.get("d2", 0.0))
                self.current_status_device = data.get("status", "")

                self.current_health_score = float(data.get("health_score", 100.0))
                self.current_trend = data.get("trend", "Mengumpulkan")
                self.current_servis = data.get("servis_estimasi", "30+ hari")
                self.current_ml_label = data.get("ml_label", "N/A")
                self.current_ml_conf = float(data.get("ml_confidence", 0.0))
                self.current_diag_label = data.get("diagnosis_label", "N/A")
                self.current_diag_conf = float(data.get("diagnosis_confidence", 0.0))
                self.current_kurtosis = float(data.get("kurtosis", 3.0))
                self.current_diagnosis_flags = data.get("diagnosis_flags", "Aman")

                self.fft_hz_buffer = data.get("fft_hz", [])
                self.fft_mag_buffer = data.get("fft_mag", [])

                self.tick += 1
                self.time_buffer.append(self.tick)
                self.v_buffer.append(self.current_v)
                self.vx_buffer.append(self.current_vx)
                self.vy_buffer.append(self.current_vy)
                self.vz_buffer.append(self.current_vz)
                self.a_buffer.append(self.current_a)
                self.temp_buffer.append(self.current_temp)
                self.rpm_buffer.append(self.current_rpm)
                self.d2_buffer.append(self.current_d2)

                self.hist_v_buf.append(self.current_v)
                self.hist_a_buf.append(self.current_a)
                self.hist_temp_buf.append(self.current_temp)
                self.hist_vx_buf.append(self.current_vx)
                self.hist_vy_buf.append(self.current_vy)
                self.hist_vz_buf.append(self.current_vz)
                self.hist_rpm_buf.append(self.current_rpm)
                self.hist_d2_buf.append(self.current_d2)

                current_st_lower = (self.current_status_device or "").strip().lower()
                if current_st_lower in ("waspada", "bahaya") and self.last_auto_snapshot_status == "normal":
                    self._trigger_auto_event_snapshot(current_st_lower)
                self.last_auto_snapshot_status = current_st_lower

                if self.recording and self.csv_writer:
                    elapsed = time.perf_counter() - self.record_start_time
                    machine_name = self.machines[self.selected_machine_idx]["name"] if self.selected_machine_idx >= 0 else "Belum Dipilih"
                    self.csv_writer.writerow([
                        round(elapsed, 3), machine_name,
                        self.current_v, self.current_a, self.current_temp,
                        self.current_vx, self.current_vy, self.current_vz,
                        self.current_rpm, self.current_d2, self.current_status_device,
                        self.current_health_score, self.current_trend, self.current_servis,
                        self.current_ml_label, self.current_diag_label, self.current_kurtosis
                    ])
                    self.csv_file.flush()
            except Exception as e:
                self.serial_connected = False
                self.ser = None
                time.sleep(1)

    def _trigger_auto_event_snapshot(self, status_severity_str):
        try:
            snapshot_filename = os.path.join(LOG_DIR, f"snapshot_{status_severity_str.upper()}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.csv")
            with open(snapshot_filename, 'w', newline='') as sf:
                sw = csv.writer(sf)
                sw.writerow([
                    'snapshot_index', 'rms_v', 'rms_a', 'temp', 'vib_x', 'vib_y', 'vib_z', 
                    'rpm', 'mahalanobis_d2', 'triggered_status'
                ])
                for idx_snap in range(len(self.hist_v_buf)):
                    sw.writerow([
                        idx_snap, self.hist_v_buf[idx_snap], self.hist_a_buf[idx_snap],
                        self.hist_temp_buf[idx_snap], self.hist_vx_buf[idx_snap],
                        self.hist_vy_buf[idx_snap], self.hist_vz_buf[idx_snap],
                        self.hist_rpm_buf[idx_snap], self.hist_d2_buf[idx_snap],
                        status_severity_str
                    ])
        except Exception as ex:
            print(f"Auto-snapshot error: {ex}")

    def _send_command(self, cmd_char):
        if self.ser is not None and self.ser.is_open:
            try:
                self.ser.write(cmd_char.encode())
            except Exception as e:
                print(f"Gagal kirim command {cmd_char}: {e}")
            
    def _apply_header_alarm_state(self, status_key):
        color = {"bahaya": COL_BAD, "waspada": COL_WARN, "diam": COL_IDLE}.get(status_key, COL_ACCENT)
        if self.packet_loss_flag:
            color = COL_WARN
        self.header_frame.setStyleSheet(f"background-color: {COL_HEADER_BG}; border-radius: 3px; border-bottom: 2px solid {color};")

    def _evaluate_diagnosis(self, v, temp, device_status=""):
        status_key = (device_status or "").strip().lower()
        self._apply_header_alarm_state(status_key)
        
        status_map = {
            "diam": ("● MOTOR DIAM", COL_IDLE),
            "bahaya": ("● DEVIASI BAHAYA", COL_BAD),
            "waspada": ("● STATUS WASPADA", COL_WARN),
            "normal": ("● SYSTEM ONLINE", COL_OK),
            "warming": ("● WARMING UP", "#888888"),
            "notcalibrated": ("● KALIBRASI BASELINE", "#888888"),
            "sensorfault": ("● SENSOR FAULT", COL_WARN),
        }
        
        txt, col = status_map.get(status_key, ("● SYSTEM ONLINE", COL_ACCENT))
        if self.packet_loss_flag:
            txt, col = "● PACKET LOSS WARNING", COL_WARN
        self.lbl_sys_status.setText(txt)
        self.lbl_sys_status.setStyleSheet(f"font-size: 7px; font-weight: bold; color: {col};")

    def _show_debug_info(self):
        port_info = self.ser.port if (self.ser is not None) else "(belum ada koneksi)"
        selected_machine_name = self.machines[self.selected_machine_idx]["name"] if self.selected_machine_idx >= 0 else "Belum dipilih"
        active_baseline = self.machines[self.selected_machine_idx].get("baseline_d2", 9.49) if self.selected_machine_idx >= 0 else "N/A"
        info = (
            f"Status koneksi   : {'TERSAMBUNG' if self.serial_connected else 'TIDAK TERSAMBUNG'}\n"
            f"Port serial      : {port_info}\n"
            f"Baud rate        : {BAUD_RATE}\n"
            f"Packet Loss Flag : {'TERDETEKSI' if self.packet_loss_flag else 'Aman'}\n"
            f"Mesin Aktif      : {selected_machine_name} (Baseline D²: {active_baseline})\n"
            f"Baris JSON akhir : {self.last_raw_line or '(belum ada data)'}\n"
            f"Vib/Snd/Tmp      : {self.current_v}, {self.current_a}, {self.current_temp}\n"
            f"RPM / D²         : {self.current_rpm}, {self.current_d2}\n"
            f"Status firmware  : {self.current_status_device or '-'}"
        )
        QMessageBox.information(self, "DEBUG - Info Koneksi & Data Terakhir", info)

    def _render_session_summary(self):
        self.lbl_session_summary.setText(
            f"Sesi: {self.session_sample_count} sample | "
            f"RPM rata-rata: {(self.session_rpm_sum / self.session_sample_count) if self.session_sample_count else 0.0:.1f} | "
            f"D² max: {self.session_d2_max:.2f} | Kondisi terparah: {self.session_worst_status}"
        )

    def _update_gui(self):
        now_dt = datetime.now()
        self.time_lbl.setText(now_dt.strftime("%H:%M:%S"))
        self.date_lbl.setText(now_dt.strftime("%d/%m/%Y"))

        self.lbl_conn_dot.setStyleSheet(
            f"font-size: 9px; font-weight: bold; color: {COL_OK if (self.serial_connected and not self.packet_loss_flag) else COL_WARN};"
        )

        if self.current_v is not None:
            self.tick += 1
            status_key = (self.current_status_device or "").strip().lower()

            # Update tampilan Mode Awam (jika aktif)
            if not self.is_technician_mode and self.selected_machine_idx >= 0:
                if status_key == "bahaya":
                    self.lbl_beranda_icon.setText("⚠️")
                    self.lbl_beranda_icon.setStyleSheet(f"font-size: 26px; color: {COL_BAD}; border: none;")
                    self.lbl_check_text.setText("BAHAYA / RISIKO KERUSAKAN")
                    self.lbl_check_text.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_BAD}; border: none;")
                elif status_key == "waspada":
                    self.lbl_beranda_icon.setText("⚡")
                    self.lbl_beranda_icon.setStyleSheet(f"font-size: 26px; color: {COL_WARN}; border: none;")
                    self.lbl_check_text.setText("WASPADA DEVIASI OPERASIONAL")
                    self.lbl_check_text.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_WARN}; border: none;")
                else:
                    self.lbl_beranda_icon.setText("✓")
                    self.lbl_beranda_icon.setStyleSheet(f"font-size: 26px; color: {COL_OK}; border: none;")
                    self.lbl_check_text.setText("MESIN AMAN / NORMAL")
                    self.lbl_check_text.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_OK}; border: none;")

                self.lbl_desc_text.setText(f"Vib: {self.current_v:.2f}G | Snd: {self.current_a:.1f}dB | Temp: {self.current_temp:.1f}°C")
                self.lbl_estimasi_servis.setText(f"Estimasi servis berikutnya: {self.current_servis}")

            # Update grafik dan data Mode Engineer / Teknisi
            self.curve_v.setData(list(self.time_buffer), list(self.v_buffer))
            self.curve_a.setData(list(self.time_buffer), list(self.a_buffer))
            self.curve_temp.setData(list(self.time_buffer), list(self.temp_buffer))

            self.curve_rpm.setData(list(self.time_buffer), list(self.rpm_buffer))
            self.curve_d2.setData(list(self.time_buffer), list(self.d2_buffer))

            if len(self.fft_hz_buffer) > 0 and len(self.fft_hz_buffer) == len(self.fft_mag_buffer):
                self.curve_fft.setData(self.fft_hz_buffer, self.fft_mag_buffer)

            self.lbl_hs_val.setText(f"{self.current_health_score:.1f}%")
            self.prog_hs.setValue(int(self.current_health_score))
            
            self.lbl_ml_val.setText(f"{self.current_ml_label} ({self.current_ml_conf*100:.0f}%)")
            self.lbl_diag_val.setText(f"{self.current_diag_label} ({self.current_diag_conf*100:.0f}%)")
            
            self.lbl_kurt_val.setText(f"{self.current_kurtosis:.2f}")
            self.lbl_flags_val.setText(f"{self.current_diagnosis_flags}")

            self.lbl_val_v.setText(f"{self.current_v:.2f} G")

            if self.current_vx is not None:
                self.lbl_val_vxyz.setText(
                    f"(X: {self.current_vx:.2f} | Y: {self.current_vy:.2f} | Z: {self.current_vz:.2f} G)"
                )
            self.lbl_val_a.setText(f"{self.current_a:.1f} dB")
            self.lbl_val_temp.setText(f"{self.current_temp:.1f} °C")
            if self.current_rpm is not None:
                self.lbl_val_rpm.setText(f"{self.current_rpm:.0f}")
            if self.current_d2 is not None:
                self.lbl_val_d2.setText(f"{self.current_d2:.2f}")

            rpm_txt = f"{self.current_rpm:.0f}" if self.current_rpm is not None else "--"
            d2_txt = f"{self.current_d2:.2f}" if self.current_d2 is not None else "--"
            self.lbl_proc_snapshot.setText(
                f"Live | Vib: {self.current_v:.2f} G | Snd: {self.current_a:.1f} dB | "
                f"Tmp: {self.current_temp:.1f} °C | RPM: {rpm_txt} | D²: {d2_txt}"
            )

            self.gauge_v.set_value(self.current_v, status_key)
            self.gauge_s.set_value(self.current_a, status_key)
            self.gauge_t.set_value(self.current_temp, status_key)

            machine_name = self.machines[self.selected_machine_idx]["name"] if self.selected_machine_idx >= 0 else "Belum dipilih"
            self.lbl_sum_machine.setText(f"Target: {machine_name}")

            self._evaluate_diagnosis(self.current_v, self.current_temp, self.current_status_device)

            if self.tick != self.last_processed_tick:
                self.last_processed_tick = self.tick
                self.session_sample_count += 1
                if self.current_rpm is not None:
                    self.session_rpm_sum += self.current_rpm
                if self.current_d2 is not None:
                    self.session_d2_max = max(self.session_d2_max, self.current_d2)

                if status_key == "waspada":
                    self.session_waspada_count += 1
                elif status_key == "bahaya":
                    self.session_bahaya_count += 1

                if status_key in STATUS_SEVERITY:
                    if STATUS_SEVERITY[status_key] > STATUS_SEVERITY.get(self.session_worst_status.lower(), 0):
                        self.session_worst_status = self.current_status_device.capitalize()

                if status_key in ("waspada", "bahaya"):
                    ts = datetime.now().strftime("%H:%M:%S")
                    event_txt = (
                        f"[{ts}] {self.current_status_device.upper()} — RPM {rpm_txt}, D² {d2_txt}, "
                        f"Vib {self.current_v:.2f}G, Tmp {self.current_temp:.1f}°C"
                    )
                    self.anomaly_events.append(event_txt)
                    if self.list_anomali.count() == 1 and self.list_anomali.item(0).text().startswith("Tidak ada"):
                        self.list_anomali.clear()
                    self.list_anomali.addItem(event_txt)
                    self.list_anomali.scrollToBottom()

                self._render_session_summary()
        elif not self.serial_connected:
            self.lbl_sys_status.setText("● MENCARI PERANGKAT (COM3)...")
            self.lbl_sys_status.setStyleSheet(f"font-size: 7px; font-weight: bold; color: {COL_WARN};")
        else:
            self.lbl_sys_status.setText("● TERSAMBUNG — MENUNGGU DATA JSON...")
            self.lbl_sys_status.setStyleSheet(f"font-size: 7px; font-weight: bold; color: {COL_ACCENT};")
            self.lbl_proc_snapshot.setText("Menunggu data serial dari ESP32...")

    def _toggle_recording(self):
        if self.selected_machine_idx < 0:
            QMessageBox.warning(self, "Perhatian", "Silakan pilih Target Mesin terlebih dahulu!")
            return
            
        if not self.recording:
            filename = os.path.join(LOG_DIR, f"log_{datetime.now().strftime('%d%m%Y_%H%M%S')}.csv")
            try:
                self.csv_file = open(filename, 'w', newline='')
                self.csv_writer = csv.writer(self.csv_file)
                self.csv_writer.writerow([
                    'timestamp', 'machine_type', 'rms_v', 'rms_a', 'temp', 'vib_x', 'vib_y', 'vib_z', 
                    'rpm', 'mahalanobis_d2', 'status', 'health_score', 'trend', 'servis_est', 'ml_label', 'diag_label', 'kurtosis'
                ])
                self.record_start_time = time.perf_counter()
                self.recording = True
                self.btn_toggle_rec.setText("BERHENTI RECORDING")
                self.btn_toggle_rec.setStyleSheet(f"background-color: {COL_BAD}; color: #ffffff; font-weight: bold; font-size: 7px; height: 20px;")
                self.lbl_rec_status.setText(f"● MENULIS -> {os.path.basename(filename)}")
            except Exception as e:
                print(f"Gagal membuat file log: {e}")
        else:
            self.recording = False
            if self.csv_file:
                self.csv_file.close()
                self.csv_file = None
            self.btn_toggle_rec.setText("MULAI RECORDING")
            self.btn_toggle_rec.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 7px; height: 20px;")
            self.lbl_rec_status.setText("● Berkas Tersimpan")
            self._refresh_log_list()

    def _refresh_log_list(self):
        self.log_list.clear()
        if os.path.isdir(LOG_DIR):
            for file in sorted(os.listdir(LOG_DIR), reverse=True):
                if file.endswith(".csv"):
                    self.log_list.addItem(file)

    def _open_log_detail(self, item):
        if item is None:
            return
        filename = item.text()
        filepath = os.path.join(LOG_DIR, filename)

        self.logdet_timer.stop()
        self.logdet_play_index = 0
        self.btn_logdet_play_pause.setText("▶ PLAY")
        self.lbl_logdet_title.setText(f"HASIL DETEKSI REKAMAN: {filename}")

        (self.logdet_times, self.logdet_v_vals, self.logdet_a_vals,
         self.logdet_t_vals) = self._logdet_load_csv(filepath)

        self.slider_logdet.setMaximum(max(0, len(self.logdet_times) - 1))
        self.slider_logdet.setValue(0)
        if self.logdet_times:
            self._logdet_render_frame(0)
        else:
            self.curve_logdet_v.clear()
            self.curve_logdet_a.clear()
            self.curve_logdet_temp.clear()
            self.lbl_logdet_pos.setText("0 / 0")
        self._logdet_render_diagnosis_summary()

        self._change_page(5)

    def _open_log_detail_from_button(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "Perhatian", "Pilih file rekaman (.csv) terlebih dahulu!")
            return
        self._open_log_detail(current_item)

    def _delete_selected_log(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            return
        filename = current_item.text()
        filepath = os.path.join(LOG_DIR, filename)
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
            self._refresh_log_list()
            self.lbl_rec_status.setText("● BERKAS DIHAPUS")
        except Exception as e:
            print(f"Gagal menghapus file: {e}")

    def _export_selected_log_to_excel(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Pilih Rekaman Dulu", "Klik salah satu berkas rekaman di daftar dulu.")
            return

        if openpyxl is None:
            QMessageBox.critical(self, "Library Belum Terinstall", "Fitur export ke Excel butuh library 'openpyxl'.")
            return

        filename = current_item.text()
        csv_path = os.path.join(LOG_DIR, filename)
        xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

        try:
            with open(csv_path, 'r') as f:
                rows = list(csv.reader(f))

            if not rows:
                QMessageBox.warning(self, "Berkas Kosong", "Berkas rekaman ini tidak memiliki data.")
                return

            header, data_rows = rows[0], rows[1:]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Sensor ESP32"

            header_fill = PatternFill(start_color="1c1e22", end_color="1c1e22", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            ws.append(header)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            status_col_idx = None
            for i, col_name in enumerate(header):
                if col_name.strip().lower() == "status":
                    status_col_idx = i
                    break

            status_fill = {
                "diam":    PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid"),
                "normal":  PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid"), 
                "waspada": PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid"),  
                "bahaya":  PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid"),   
            }

            for row in data_rows:
                converted_row = []
                for value in row:
                    try:
                        converted_row.append(float(value))
                    except (ValueError, TypeError):
                        converted_row.append(value)
                ws.append(converted_row)

                if status_col_idx is not None:
                    status_val = str(row[status_col_idx]).strip().lower()
                    fill = status_fill.get(status_val)
                    if fill:
                        ws.cell(row=ws.max_row, column=status_col_idx + 1).fill = fill

            for col_cells in ws.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

            ws.freeze_panes = "A2"  
            wb.save(xlsx_path)

            QMessageBox.information(
                self, "Export Berhasil",
                f"Data berhasil diexport ke:\n{xlsx_path}\n\nTotal {len(data_rows)} baris data."
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Gagal", f"Terjadi kesalahan saat export:\n{e}")

class GlobalEscHandler(QApplication):
    def notify(self, receiver, event):
        if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Escape:
            for widget in self.topLevelWidgets():
                if isinstance(widget, Dashboard):
                    if widget.csv_file:
                        widget.csv_file.close()
                    widget.close()
                    return True
        return super().notify(receiver, event)

if __name__ == '__main__':
    app = GlobalEscHandler(sys.argv)
    db = Dashboard()
    db.show()
    db.raise_()
    db.activateWindow()
    sys.exit(app.exec_())
